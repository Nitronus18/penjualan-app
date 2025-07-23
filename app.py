# app.py

from flask import Flask, render_template, request, redirect, url_for, flash, Response
from datetime import datetime, timedelta, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from zoneinfo import ZoneInfo
import sqlite3
import os
import csv
import io
from io import BytesIO
import webbrowser
import threading
import logging
import uuid

datetime.now(timezone(timedelta(hours=7)))
WIB = timezone(timedelta(hours=7))

print("📁 Lokasi database saat ini:", os.path.abspath('database.db'))

logging.basicConfig(
    filename='log.txt',
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
def buat_database_dan_seed():

    conn = sqlite3.connect('database.db')
    cur = conn.cursor()

    # Buat tabel produk
    cur.execute('''
    CREATE TABLE IF NOT EXISTS produk (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nama TEXT NOT NULL,
        harga INTEGER NOT NULL,
        stok INTEGER NOT NULL,
        kategori TEXT
    );
    ''')

    # Buat tabel penjualan
    cur.execute('''
    CREATE TABLE IF NOT EXISTS penjualan (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        produk_id INTEGER,
        jumlah INTEGER,
        tanggal TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (produk_id) REFERENCES produk(id)
    );
    ''')

   # ➕ Tambahkan kolom transaksi_id jika belum ada
    cur.execute("PRAGMA table_info(penjualan)")
    kolom_penjualan = [row[1] for row in cur.fetchall()]
    if 'transaksi_id' not in kolom_penjualan:
        print("🛠️ Menambahkan kolom transaksi_id ke tabel penjualan...")
        cur.execute("ALTER TABLE penjualan ADD COLUMN transaksi_id TEXT")    

    cur.execute('''
    CREATE TABLE IF NOT EXISTS bahan (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nama TEXT NOT NULL,
        satuan TEXT
    );
    ''')

    cur.execute('''
    CREATE TABLE IF NOT EXISTS pembelian (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        bahan_id INTEGER,
        jumlah REAL,
        harga_total INTEGER,
        tanggal TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (bahan_id) REFERENCES bahan(id)
    );
    ''')
    cur.execute('''
    CREATE TABLE IF NOT EXISTS bahan (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nama TEXT NOT NULL,
        satuan TEXT
    )
    ''')
    # Buat penjualan_warnet jika belum ada
    cur.execute('''
    CREATE TABLE IF NOT EXISTS penjualan_warnet (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        bahan_id INTEGER,
        jumlah REAL,
        harga_total INTEGER,
        tanggal TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (bahan_id) REFERENCES bahan(id)
    )
    ''')       

    # Daftar produk tetap
    produk_list = [
    # Chocolate Series
    ("coco rocky tiramisu", 8000, 0, "coklat"),
    ("coco banana", 8000, 0, "coklat"),
    ("coco classic", 8000, 0, "coklat"),
    ("coco halloween", 8000, 0, "coklat"),
    ("coco kulaku", 8000, 0, "coklat"),
    ("coco lumer", 8000, 0, "coklat"),
    ("coco maltino", 8000, 0, "coklat"),
    ("coco red florest", 8000, 0, "coklat"),
    ("coco trouble", 8000, 0, "coklat"),

    # Fruit Series
    ("alpukat", 8000, 0, "fruit"),
    ("mango", 8000, 0, "fruit"),
    ("lychee", 8000, 0, "fruit"),
    ("strawberry", 8000, 0, "fruit"),
    ("blueberry", 8000, 0, "fruit"),
    ("durian", 8000, 0, "fruit"),
    ("blewah", 8000, 0, "fruit"),
    ("nanas", 8000, 0, "fruit"),

    # Spesial Series
    ("brown sugar tiramisu", 8000, 0, "spesial"),
    ("brown sugar vanilla lagoan", 8000, 0, "spesial"),
    ("blody velvet", 8000, 0, "spesial"),
    ("brows sugar durian", 8000, 0, "spesial"),
    ("brown sugar klepon", 8000, 0, "spesial"),
    ("green velvet", 8000, 0, "spesial"),
    ("violet", 8000, 0, "spesial"),

    # Coffee Series
    ("coffe ala goklat", 8000, 0, "coffee"),
    ("coffe tiramisu", 8000, 0, "coffee"),
    ("coconut aren coffe", 8000, 0, "coffee"),
    ("capucino kulaku", 8000, 0, "coffee"),
    ("durian coffe", 8000, 0, "coffee"),
    ("coffe pandan", 8000, 0, "coffee"),
        # Gotea Series
    ("gotea honey", 6000, 0, "gotea"),
    ("gotea jasmine", 6000, 0, "gotea"),
    ("gotea special", 6000, 0, "gotea"),
    ("gotea black current", 7000, 0, "gotea"),
    ("gotea apple", 7000, 0, "gotea"),
    ("gotea mangga", 7000, 0, "gotea"),
    ("gotea lychee", 7000, 0, "gotea"),
    ("gotea lemon", 7000, 0, "gotea"),
    ("gotea peach", 7000, 0, "gotea"),
    ("gotea grape", 7000, 0, "gotea"),

    #Es Teh Jumbo
    ("EsTeh Jumbo", 3000, 0, "EsTeh"),
    ("EsTeh Jumbo Jeruk Nipis", 4000, 0, "EsTeh"),
]

    for nama, harga, stok, kategori in produk_list:
        cur.execute("SELECT COUNT(*) FROM produk WHERE nama = ?", (nama,))
        if cur.fetchone()[0] == 0:
            cur.execute("INSERT INTO produk (nama, harga, stok, kategori) VALUES (?, ?, ?, ?)", (nama, harga, stok, kategori))
        else:
            cur.execute("UPDATE produk SET harga = ?, kategori = ? WHERE nama = ?", (harga, kategori, nama))

    conn.commit()
    conn.close()

def cek_dan_buat_database():
    buat = False
    if not os.path.exists('database.db'):
        buat = True
    else:
        try:
            conn = sqlite3.connect('database.db')
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM produk")
            conn.close()
        except sqlite3.OperationalError:
            buat = True

    if buat:
        print("📦 Membuat database dan isi produk...")
        buat_database_dan_seed()

def redirect_back(default='index'):
    next_url = request.form.get('next') or request.args.get('next')
    if next_url:
        return redirect(next_url)
    return redirect(url_for(default))


# Panggil sebelum Flask start
cek_dan_buat_database()



waktu_lokal = datetime.now()

app = Flask(__name__)
app.secret_key = 'sehatidotcom'  # kunci bebas, wajib untuk flash

@app.template_filter('rupiah')
def format_rupiah(value):
    return "Rp{:,.0f}".format(value).replace(",", ".")

def get_db_connection():
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row
    return conn

@app.route('/')
def index():
    sort = request.args.get('sort', 'stok')
    query = request.args.get('q', '').strip().lower()

    conn = get_db_connection()

    if query:
        produk = conn.execute(f'''
            SELECT * FROM produk
            WHERE LOWER(nama) LIKE ?
            ORDER BY {sort} ASC
        ''', (f'%{query}%',)).fetchall()
    else:
        produk = conn.execute(f'''
            SELECT * FROM produk ORDER BY {sort} ASC
        ''').fetchall()

    penjualan = conn.execute('''
        SELECT p.id, pr.nama, p.jumlah, pr.harga, p.tanggal
        FROM penjualan p
        JOIN produk pr ON pr.id = p.produk_id
        ORDER BY p.tanggal DESC
    ''').fetchall()

    total_pemasukan = sum(p['jumlah'] * p['harga'] for p in penjualan)
    conn.close()
    return render_template('index.html', produk=produk, penjualan=penjualan,
                           total=total_pemasukan, current_sort=sort, query=query,
                           active_page='home')



@app.route('/tambah-produk', methods=['POST'])
def tambah_produk():
    nama = request.form['nama']
    harga = int(request.form['harga'])
    stok = int(request.form['stok'])
    kategori = request.form['kategori']

    conn = get_db_connection()
    conn.execute('INSERT INTO produk (nama, harga, stok, kategori) VALUES (?, ?, ?, ?)',
                 (nama, harga, stok, kategori))
    conn.commit()
    conn.close()
    flash('✅ Produk berhasil ditambahkan.')
    return redirect(url_for('index'))


@app.route('/jual', methods=['POST'])
def jual():
    produk_id = int(request.form['produk_id'])
    jumlah = int(request.form['jumlah'])

    conn = get_db_connection()
    produk = conn.execute('SELECT stok, harga FROM produk WHERE id = ?', (produk_id,)).fetchone()

    if produk and produk['stok'] >= jumlah:
        conn.execute('UPDATE produk SET stok = stok - ? WHERE id = ?', (jumlah, produk_id))

        waktu_wib = datetime.now(WIB).strftime('%Y-%m-%d %H:%M:%S')

        try:
            conn.execute('INSERT INTO penjualan (produk_id, jumlah, tanggal) VALUES (?, ?, ?)',
                         (produk_id, jumlah, waktu_wib))
            conn.commit()
            logging.info(f"Penjualan dicatat: produk_id={produk_id}, jumlah={jumlah}, waktu={waktu_wib}")
        except Exception as e:
            logging.error(f"Gagal simpan penjualan: {e}")
    else:
        logging.warning(f"Penjualan gagal: stok kurang / produk tidak ditemukan - produk_id={produk_id}")

    conn.close()
    return redirect_back()

@app.route('/jual-multi', methods=['POST'])
def jual_multi():
    raw_data = request.form.get('items', '')
    items = raw_data.split(',') if raw_data else []

    conn = get_db_connection()
    waktu = datetime.now(WIB).strftime('%Y-%m-%d %H:%M:%S')
    transaksi_id = str(uuid.uuid4())  # 👉 unik untuk satu transaksi

    for item in items:
        try:
            produk_id, jumlah = map(int, item.split(':'))
            produk = conn.execute('SELECT stok FROM produk WHERE id = ?', (produk_id,)).fetchone()
            if produk and produk['stok'] >= jumlah:
                conn.execute('UPDATE produk SET stok = stok - ? WHERE id = ?', (jumlah, produk_id))
                conn.execute('''
                    INSERT INTO penjualan (produk_id, jumlah, tanggal, transaksi_id)
                    VALUES (?, ?, ?, ?)
                ''', (produk_id, jumlah, waktu, transaksi_id))
        except Exception as e:
            print(f"❌ ERROR: {e}")
            continue

    conn.commit()
    conn.close()
    flash("✅ Transaksi berhasil ditambahkan.")
    return redirect(url_for('produk_kategori', kategori=request.args.get('kategori', 'coklat'), transaksi='ok'))







@app.route('/hapus-penjualan/<int:id>', methods=['POST'])
def hapus_penjualan(id):
    conn = get_db_connection()

    # Ambil data penjualan (produk_id & jumlah)
    penjualan = conn.execute('SELECT produk_id, jumlah FROM penjualan WHERE id = ?', (id,)).fetchone()

    if penjualan:
        # Kembalikan stok produk
        conn.execute('UPDATE produk SET stok = stok + ? WHERE id = ?', (penjualan['jumlah'], penjualan['produk_id']))
        # Hapus transaksi
        conn.execute('DELETE FROM penjualan WHERE id = ?', (id,))
        conn.commit()

    conn.close()
    return redirect(url_for('halaman_penjualan'))


@app.route('/produk/tersedia')
def produk_tersedia():
    conn = get_db_connection()
    produk = conn.execute('SELECT * FROM produk WHERE stok > 0 ORDER BY nama ASC').fetchall()
    conn.close()
    return render_template('produk_by_stok.html', produk=produk, judul="📦 Produk Tersedia")

@app.route('/produk/kosong')
def produk_kosong():
    conn = get_db_connection()
    produk = conn.execute('SELECT * FROM produk WHERE stok = 0 ORDER BY nama ASC').fetchall()
    conn.close()
    return render_template('produk_by_stok.html', produk=produk, judul="❌ Produk Stok Habis")


@app.route('/update-stok/<int:id>', methods=['POST'])
def update_stok(id):
    stok_baru = int(request.form['stok'])

    conn = get_db_connection()
    conn.execute('UPDATE produk SET stok = ? WHERE id = ?', (stok_baru, id))
    conn.commit()
    conn.close()

    flash('✅ Stok berhasil diperbarui.')
    return redirect_back()



@app.route('/produk/<kategori>')
def produk_kategori(kategori):
    query = request.args.get('q', '').strip().lower()
    stok_filter = request.args.get('stok', 'semua')

    conn = get_db_connection()

    base_query = 'SELECT * FROM produk WHERE kategori = ?'
    params = [kategori]

    if stok_filter == 'tersedia':
        base_query += ' AND stok > 0'
    elif stok_filter == 'habis':
        base_query += ' AND stok = 0'

    if query:
        base_query += ' AND LOWER(nama) LIKE ?'
        params.append(f'%{query}%')

    base_query += ' ORDER BY nama ASC'

    produk = conn.execute(base_query, params).fetchall()
    conn.close()

    kategori_label = kategori.capitalize()
    return render_template('produk_by_kategori.html', produk=produk,
                           kategori=kategori_label, query=query, stok_filter=stok_filter,  active_page='produk-' + kategori)

@app.route('/produk')
def semua_produk():
    query = request.args.get('q', '').strip().lower()
    stok_filter = request.args.get('stok', 'semua')

    conn = get_db_connection()
    base_query = 'SELECT * FROM produk WHERE 1=1'
    params = []

    if stok_filter == 'tersedia':
        base_query += ' AND stok > 0'
    elif stok_filter == 'habis':
        base_query += ' AND stok = 0'

    if query:
        base_query += ' AND LOWER(nama) LIKE ?'
        params.append(f'%{query}%')

    base_query += ' ORDER BY nama ASC'

    produk = conn.execute(base_query, params).fetchall()
    conn.close()

    return render_template('produk_by_kategori.html', produk=produk,
                           kategori="Semua Produk", query=query, stok_filter=stok_filter,
                           active_page='produk')



@app.route('/laporan')
def laporan():
    mode = request.args.get('mode', 'harian')
    tanggal = request.args.get('tanggal', datetime.now().strftime('%Y-%m-%d'))

    conn = get_db_connection()
    cursor = conn.cursor()

    labels = []
    values = []

    if mode == 'harian':
        cursor.execute('''
            SELECT DATE(tanggal) AS hari, SUM(p.jumlah * pr.harga) AS total
            FROM penjualan p
            JOIN produk pr ON pr.id = p.produk_id
            WHERE DATE(tanggal) = ?
            GROUP BY hari
        ''', (tanggal,))
        for row in cursor.fetchall():
            labels.append(row['hari'])
            values.append(row['total'] or 0)

    elif mode == 'mingguan':
        start = datetime.strptime(tanggal, "%Y-%m-%d")
        dates = [(start - timedelta(days=i)).date().isoformat() for i in reversed(range(7))]
        for d in dates:
            cursor.execute('''
                SELECT SUM(p.jumlah * pr.harga) AS total
                FROM penjualan p
                JOIN produk pr ON pr.id = p.produk_id
                WHERE DATE(tanggal) = ?
            ''', (d,))
            total = cursor.fetchone()['total'] or 0
            labels.append(d)
            values.append(total)

    elif mode == 'bulanan':
        bulan = tanggal[:7]  # 'YYYY-MM'
        cursor.execute('''
            SELECT DATE(tanggal) AS hari, SUM(p.jumlah * pr.harga) AS total
            FROM penjualan p
            JOIN produk pr ON pr.id = p.produk_id
            WHERE strftime('%Y-%m', tanggal) = ?
            GROUP BY hari
            ORDER BY hari
        ''', (bulan,))
        for row in cursor.fetchall():
            labels.append(row['hari'])
            values.append(row['total'] or 0)

    produk = conn.execute('SELECT * FROM produk ORDER BY stok ASC').fetchall()
    conn.close()
    return render_template('laporan.html',
                       produk=produk,
                       labels=labels,
                       values=values,
                       mode=mode,
                       tanggal=tanggal,
                       active_page='laporan')



@app.route('/penjualan')
def halaman_penjualan():
    filter_tanggal = request.args.get('tanggal')
    filter_kategori = request.args.get('kategori', 'semua')

    conn = get_db_connection()

    base_query = '''
        SELECT p.transaksi_id, p.tanggal, pr.nama, pr.kategori, p.jumlah, pr.harga
        FROM penjualan p
        JOIN produk pr ON pr.id = p.produk_id
    '''
    where_clauses = []
    params = []

    if filter_tanggal:
        where_clauses.append('DATE(p.tanggal) = ?')
        params.append(filter_tanggal)
    if filter_kategori != 'semua':
        where_clauses.append('pr.kategori = ?')
        params.append(filter_kategori)

    if where_clauses:
        base_query += ' WHERE ' + ' AND '.join(where_clauses)

    base_query += ' ORDER BY p.tanggal DESC'

    raw = conn.execute(base_query, params).fetchall()
    conn.close()
    # Hitung total berdasarkan produk (bukan transaksi)
    total = 0
    total_goklat = 0
    total_esteh = 0

    for row in raw:
        subtotal = row['jumlah'] * row['harga']
        total += subtotal
        if row['kategori'] == 'EsTeh':
            total_esteh += subtotal
        else:
            total_goklat += subtotal

    # Gabungkan per transaksi_id
    grouped = {}
    for row in raw:
        tid = row['transaksi_id'] or f"single-{row['tanggal']}-{row['nama']}"
        if tid not in grouped:
            grouped[tid] = {
                'tanggal': row['tanggal'],
                'items': [],
                'kategori': set(),
                'total': 0
            }
        grouped[tid]['items'].append(f"{row['nama']} x{row['jumlah']}")
        grouped[tid]['kategori'].add(row['kategori'])
        grouped[tid]['total'] += row['jumlah'] * row['harga']

    penjualan = [{
        'id': tid,
        'tanggal': data['tanggal'],
        'deskripsi': ', '.join(data['items']),
        'kategori': ', '.join(data['kategori']),
        'total': data['total']
    } for tid, data in grouped.items()]


    return render_template('penjualan.html',
        penjualan=penjualan,
        total=total,
        total_goklat=total_goklat,
        total_esteh=total_esteh,
        filter_tanggal=filter_tanggal,
        filter_kategori=filter_kategori,
        active_page='penjualan')


@app.route('/hapus-transaksi/<transaksi_id>', methods=['POST'])
def hapus_transaksi(transaksi_id):
    conn = get_db_connection()

    # Ambil semua item transaksi
    items = conn.execute('SELECT produk_id, jumlah FROM penjualan WHERE transaksi_id = ?', (transaksi_id,)).fetchall()
    for item in items:
        conn.execute('UPDATE produk SET stok = stok + ? WHERE id = ?', (item['jumlah'], item['produk_id']))

    conn.execute('DELETE FROM penjualan WHERE transaksi_id = ?', (transaksi_id,))
    conn.commit()
    conn.close()
    flash("✅ Transaksi berhasil dihapus.")
    return redirect(url_for('halaman_penjualan'))


@app.route('/export/laporan')
def export_laporan_csv():
    conn = get_db_connection()
    produk = conn.execute('SELECT * FROM produk ORDER BY stok ASC').fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow(['Nama Produk', 'Harga Satuan', 'Stok', 'Total Nilai'])

    for p in produk:
        writer.writerow([p['nama'], p['harga'], p['stok'], p['harga'] * p['stok']])

    response = Response(output.getvalue(), mimetype='text/csv')
    response.headers['Content-Disposition'] = 'attachment; filename=laporan_stok.csv'
    return response


@app.route('/export/laporan-gabungan-xlsx')
def export_laporan_gabungan_xlsx():
    mode = request.args.get('mode', 'harian')
    tanggal_str = request.args.get('tanggal', datetime.now().strftime('%Y-%m-%d'))

    try:
        tanggal = datetime.strptime(tanggal_str, '%Y-%m-%d')
    except ValueError:
        return "❌ Format tanggal tidak valid. Gunakan YYYY-MM-DD", 400

    # Tentukan rentang tanggal
    if mode == 'harian':
        tanggal_list = [tanggal.date()]
    elif mode == 'mingguan':
        start = (tanggal - timedelta(days=tanggal.weekday())).date()
        tanggal_list = [start + timedelta(days=i) for i in range(7)]
    elif mode == 'bulanan':
        start = tanggal.replace(day=1).date()
        next_month = (tanggal.replace(day=28) + timedelta(days=4)).replace(day=1)
        end = (next_month - timedelta(days=1)).date()
        tanggal_list = [start + timedelta(days=i) for i in range((end - start).days + 1)]
    else:
        return "❌ Mode tidak dikenali (gunakan harian, mingguan, bulanan)", 400

    conn = get_db_connection()
    cursor = conn.cursor()

    wb = Workbook()
    wb.remove(wb.active)

    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4B2C20")
    total_fill = PatternFill("solid", fgColor="D5BBA2")
    center_align = Alignment(horizontal="center")

    def write_table(ws, title, headers, rows, total_column_idx=None):
        ws.append([title])
        row = ws.max_row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = center_align

        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        total = 0
        for r in rows:
            ws.append(r)
            if total_column_idx is not None:
                total += r[total_column_idx]

        if total_column_idx is not None:
            ws.append([''] * (total_column_idx) + ['TOTAL', total])
            for col in range(total_column_idx + 1, total_column_idx + 3):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.font = Font(bold=True)
                cell.fill = total_fill
        ws.append([])

    # Loop per tanggal
    for tgl in tanggal_list:
        sheet = wb.create_sheet(title=tgl.strftime('%Y-%m-%d'))

        # Penjualan
        penjualan = cursor.execute('''
            SELECT pr.nama, pr.harga, pr.kategori, SUM(p.jumlah) as jumlah
            FROM penjualan p
            JOIN produk pr ON pr.id = p.produk_id
            WHERE DATE(p.tanggal) = ?
            GROUP BY pr.id
            ORDER BY pr.nama ASC
        ''', (tgl.isoformat(),)).fetchall()

        goklat_rows, esteh_rows = [], []
        no_goklat, no_esteh = 1, 1

        for p in penjualan:
            total = (p['jumlah'] or 0) * p['harga']
            row = [None, p['nama'], p['harga'], p['jumlah'], total]

            if p['kategori'] == 'EsTeh':
                row[0] = no_esteh
                esteh_rows.append(row)
                no_esteh += 1
            else:
                row[0] = no_goklat
                goklat_rows.append(row)
                no_goklat += 1

        write_table(sheet, "PENJUALAN GOKLAT", ['No', 'Varian', 'Harga', 'Jumlah', 'Total'], goklat_rows, total_column_idx=4)
        write_table(sheet, "PENJUALAN ESTEHT", ['No', 'Varian', 'Harga', 'Jumlah', 'Total'], esteh_rows, total_column_idx=4)

        # Pembelian
        pembelian = cursor.execute('''
            SELECT b.nama, b.satuan, p.jumlah, p.harga_total
            FROM pembelian p
            JOIN bahan b ON b.id = p.bahan_id
            WHERE DATE(p.tanggal) = ?
            ORDER BY b.nama ASC
        ''', (tgl.isoformat(),)).fetchall()

        pembelian_rows = [[idx+1, p['nama'], p['jumlah'], p['satuan'], p['harga_total']] for idx, p in enumerate(pembelian)]
        write_table(sheet, "PEMBELIAN BAHAN", ['No', 'Nama Bahan', 'Jumlah', 'Satuan', 'Total'], pembelian_rows, total_column_idx=4)
    # Penjualan Warnet
    warnet = cursor.execute('''
        SELECT b.nama, b.satuan, p.jumlah, p.harga_total
        FROM penjualan_warnet p
        JOIN bahan b ON b.id = p.bahan_id
        WHERE DATE(p.tanggal) = ?
        ORDER BY b.nama ASC
    ''', (tgl.isoformat(),)).fetchall()

    warnet_rows = [[idx+1, w['nama'], w['jumlah'], w['satuan'], w['harga_total']] for idx, w in enumerate(warnet)]
    write_table(sheet, "PENJUALAN WARNET", ['No', 'Nama', 'Jumlah', 'Satuan', 'Total'], warnet_rows, total_column_idx=4)

    conn.close()

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"laporan_gabungan_{mode}_{tanggal_str}.xlsx"
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={filename}"}
    )


@app.route('/warnet', methods=['GET', 'POST'])
def penjualan_warnet():
    conn = get_db_connection()

    if request.method == 'POST':
        bahan_id = int(request.form['bahan_id'])
        jumlah = float(request.form['jumlah'])
        harga_total = int(request.form['harga_total'])
        waktu_wib = datetime.now(WIB).strftime('%Y-%m-%d %H:%M:%S')
        conn.execute('INSERT INTO penjualan_warnet (bahan_id, jumlah, harga_total, tanggal) VALUES (?, ?, ?, ?)',
                     (bahan_id, jumlah, harga_total, waktu_wib))
        conn.commit()
        flash("✅ Penjualan warnet berhasil dicatat.")
        return redirect(url_for('penjualan_warnet'))

    mode = request.args.get('mode', 'harian')
    tanggal = request.args.get('tanggal', datetime.now(WIB).strftime('%Y-%m-%d'))

    # ambil data berdasarkan filter waktu
    # ...

    bahan_list = conn.execute('SELECT * FROM bahan ORDER BY nama ASC').fetchall()

    # ambil data penjualan_warnet dengan join bahan
    cursor = conn.cursor()
    data = cursor.execute('''
        SELECT p.id, b.nama, p.jumlah, p.harga_total, p.tanggal
        FROM penjualan_warnet p
        JOIN bahan b ON b.id = p.bahan_id
        WHERE DATE(p.tanggal) = ?
        ORDER BY p.tanggal DESC
    ''', (tanggal,)).fetchall()

    mode = request.args.get('mode', 'harian')
    tanggal = request.args.get('tanggal', datetime.now(WIB).strftime('%Y-%m-%d'))

    cursor = conn.cursor()
    params = []
    where_clause = ""

    if mode == 'harian':
        where_clause = "WHERE DATE(p.tanggal) = ?"
        params.append(tanggal)

    elif mode == 'mingguan':
        ref = datetime.strptime(tanggal, "%Y-%m-%d")
        start = (ref - timedelta(days=ref.weekday())).date()
        end = start + timedelta(days=6)
        where_clause = "WHERE DATE(p.tanggal) BETWEEN ? AND ?"
        params.extend([start.isoformat(), end.isoformat()])

    elif mode == 'bulanan':
        bulan = tanggal[:7]
        where_clause = "WHERE strftime('%Y-%m', p.tanggal) = ?"
        params.append(bulan)

    query = f'''
        SELECT p.id, b.nama, p.jumlah, p.harga_total, p.tanggal
        FROM penjualan_warnet p
        JOIN bahan b ON b.id = p.bahan_id
        {where_clause}
        ORDER BY p.tanggal DESC
    '''

    data = cursor.execute(query, params).fetchall()
    total_pendapatan = sum(d['harga_total'] for d in data)
        

    total_pendapatan = sum(d['harga_total'] for d in data)
    conn.close()

    return render_template('penjualan_warnet.html',
                           bahan=bahan_list,
                           data=data,
                           total_pendapatan=total_pendapatan,
                           mode=mode,
                           tanggal=tanggal,
                           active_page='warnet')



@app.route('/export/penjualan-warnet-xlsx')
def export_penjualan_warnet_xlsx():
    mode = request.args.get('mode', 'harian')
    tanggal_str = request.args.get('tanggal', datetime.now().strftime('%Y-%m-%d'))

    try:
        tanggal = datetime.strptime(tanggal_str, '%Y-%m-%d')
    except ValueError:
        return "❌ Format tanggal tidak valid.", 400

    # Hitung rentang tanggal
    if mode == 'harian':
        tanggal_list = [tanggal.date()]
    elif mode == 'mingguan':
        start = (tanggal - timedelta(days=tanggal.weekday())).date()
        tanggal_list = [start + timedelta(days=i) for i in range(7)]
    elif mode == 'bulanan':
        start = tanggal.replace(day=1).date()
        next_month = (tanggal.replace(day=28) + timedelta(days=4)).replace(day=1)
        end = (next_month - timedelta(days=1)).date()
        tanggal_list = [start + timedelta(days=i) for i in range((end - start).days + 1)]
    else:
        return "❌ Mode tidak dikenali", 400

    conn = get_db_connection()
    cursor = conn.cursor()

    wb = Workbook()
    wb.remove(wb.active)

    # Style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4B2C20")
    total_fill = PatternFill("solid", fgColor="D5BBA2")
    center = Alignment(horizontal="center")

    for tgl in tanggal_list:
        ws = wb.create_sheet(title=tgl.strftime('%Y-%m-%d'))

        rows = cursor.execute('''
            SELECT b.nama, b.satuan, p.jumlah, p.harga_total
            FROM penjualan_warnet p
            JOIN bahan b ON b.id = p.bahan_id
            WHERE DATE(p.tanggal) = ?
            ORDER BY b.nama ASC
        ''', (tgl.isoformat(),)).fetchall()

        ws.append(['PENJUALAN WARNET'])
        r = ws.max_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=1).alignment = center

        ws.append(['No', 'Jenis', 'Jumlah', 'Satuan', 'Total (Rp)'])
        for col in range(1, 6):
            c = ws.cell(row=ws.max_row, column=col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center

        total = 0
        for idx, row in enumerate(rows, start=1):
            ws.append([idx, row['nama'], row['jumlah'], row['satuan'], row['harga_total']])
            total += row['harga_total']

        ws.append(['', '', '', 'TOTAL', total])
        for col in range(4, 6):
            c = ws.cell(row=ws.max_row, column=col)
            c.font = Font(bold=True)
            c.fill = total_fill

    conn.close()
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"penjualan_warnet_{mode}_{tanggal_str}.xlsx"
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={filename}"}
    )


@app.route('/stok/update')
def update_semua_stok():
    kategori = request.args.get('kategori', 'semua')

    conn = get_db_connection()
    if kategori == 'semua':
        produk = conn.execute('SELECT * FROM produk ORDER BY nama ASC').fetchall()
    else:
        produk = conn.execute('SELECT * FROM produk WHERE kategori = ? ORDER BY nama ASC', (kategori,)).fetchall()
    conn.close()

    return render_template('update_stok.html',
                       produk=produk,
                       kategori_terpilih=kategori,
                       active_page='stok')


    



@app.route('/export/penjualan-ringkas')
def export_penjualan_ringkas():
    conn = get_db_connection()
    cursor = conn.cursor()

    produk = cursor.execute('SELECT id, nama, harga, kategori FROM produk ORDER BY nama ASC').fetchall()
    penjualan = cursor.execute('''
        SELECT produk_id, SUM(jumlah) as total_jumlah
        FROM penjualan
        GROUP BY produk_id
    ''').fetchall()
    conn.close()

    penjualan_dict = {p['produk_id']: p['total_jumlah'] for p in penjualan}

    goklat_rows = []
    esteh_rows = []
    total_goklat = 0
    total_esteh = 0
    no_goklat = 1
    no_esteh = 1

    for p in produk:
        jumlah = penjualan_dict.get(p['id'], 0)
        total = jumlah * p['harga']

        row = [None, p['nama'], p['harga'], jumlah, total]

        if p['kategori'] == 'EsTeh':
            row[0] = no_esteh
            esteh_rows.append(row)
            total_esteh += total
            no_esteh += 1
        else:
            row[0] = no_goklat
            goklat_rows.append(row)
            total_goklat += total
            no_goklat += 1

    # Buat CSV
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')

    # GOKLAT SECTION
    writer.writerow(['=== PENJUALAN GOKLAT ==='])
    writer.writerow(['No', 'Varian', 'Harga', 'Jumlah', 'Total'])
    for row in goklat_rows:
        writer.writerow(row)
    writer.writerow(['', '', '', 'TOTAL', total_goklat])
    writer.writerow([])

    # ESTEHT SECTION
    writer.writerow(['=== PENJUALAN ESTEHT ==='])
    writer.writerow(['No', 'Varian', 'Harga', 'Jumlah', 'Total'])
    for row in esteh_rows:
        writer.writerow(row)
    writer.writerow(['', '', '', 'TOTAL', total_esteh])

    response = Response(output.getvalue(), mimetype='text/csv')
    response.headers['Content-Disposition'] = 'attachment; filename=penjualan_ringkas.csv'
    return response

@app.route('/export/penjualan-ringkas-xlsx')
def export_penjualan_ringkas_xlsx():
    mode = request.args.get('mode', 'harian')
    tanggal_str = request.args.get('tanggal', datetime.now().strftime('%Y-%m-%d'))

    try:
        tanggal = datetime.strptime(tanggal_str, '%Y-%m-%d')
    except ValueError:
        return "❌ Format tanggal tidak valid. Gunakan YYYY-MM-DD", 400

    # Hitung rentang tanggal
    if mode == 'harian':
        tanggal_list = [tanggal.date()]
    elif mode == 'mingguan':
        start = (tanggal - timedelta(days=tanggal.weekday())).date()
        tanggal_list = [start + timedelta(days=i) for i in range(7)]
    elif mode == 'bulanan':
        start = tanggal.replace(day=1).date()
        next_month = (tanggal.replace(day=28) + timedelta(days=4)).replace(day=1)
        end = (next_month - timedelta(days=1)).date()
        tanggal_list = [start + timedelta(days=i) for i in range((end - start).days + 1)]
    else:
        return "❌ Mode tidak dikenali (gunakan harian, mingguan, bulanan)", 400

    # DB & Workbook
    conn = get_db_connection()
    cursor = conn.cursor()
    wb = Workbook()
    wb.remove(wb.active)

    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4B2C20")
    total_fill = PatternFill("solid", fgColor="D5BBA2")
    center_align = Alignment(horizontal="center")

    def write_section(ws, title, data_rows, total):
        ws.append([title])
        row = ws.max_row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = center_align

        ws.append(['No', 'Varian', 'Harga', 'Jumlah', 'Total'])
        for col in range(1, 6):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for r in data_rows:
            ws.append(r)

        ws.append(['', '', '', 'TOTAL', total])
        for col in range(4, 6):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = total_fill

        ws.append([])

    # Loop per tanggal → per sheet
    for tgl in tanggal_list:
        sheet = wb.create_sheet(title=tgl.strftime('%Y-%m-%d'))

        query = '''
            SELECT pr.id, pr.nama, pr.harga, pr.kategori, SUM(p.jumlah) as total_jumlah
            FROM penjualan p
            JOIN produk pr ON pr.id = p.produk_id
            WHERE DATE(p.tanggal) = ?
            GROUP BY pr.id, pr.nama, pr.harga, pr.kategori
            ORDER BY pr.nama ASC
        '''
        penjualan = cursor.execute(query, (tgl.isoformat(),)).fetchall()

        goklat_rows, esteh_rows = [], []
        no_goklat, no_esteh = 1, 1
        total_goklat, total_esteh = 0, 0

        for p in penjualan:
            jumlah = p['total_jumlah'] or 0
            total = jumlah * p['harga']
            row = [None, p['nama'], p['harga'], jumlah, total]

            if p['kategori'] == 'EsTeh':
                row[0] = no_esteh
                esteh_rows.append(row)
                total_esteh += total
                no_esteh += 1
            else:
                row[0] = no_goklat
                goklat_rows.append(row)
                total_goklat += total
                no_goklat += 1

        write_section(sheet, "PENJUALAN GOKLAT", goklat_rows, total_goklat)
        write_section(sheet, "PENJUALAN ESTEHT", esteh_rows, total_esteh)

    conn.close()

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"penjualan_{mode}_{tanggal_str}_perhari.xlsx"
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={filename}"}
    )



@app.route('/pembelian', methods=['GET', 'POST'])
def pembelian():
    conn = get_db_connection()

    if request.method == 'POST':
        bahan_id = int(request.form['bahan_id'])
        jumlah = float(request.form['jumlah'])
        harga_total = int(request.form['harga_total'])

        waktu_wib = datetime.now(WIB).strftime('%Y-%m-%d %H:%M:%S')
        conn.execute('INSERT INTO pembelian (bahan_id, jumlah, harga_total, tanggal) VALUES (?, ?, ?, ?)',
                     (bahan_id, jumlah, harga_total, waktu_wib))
        conn.commit()
        conn.close()
        flash('✅ Pembelian berhasil dicatat.')
        return redirect(url_for('pembelian'))

    # Filter
    mode = request.args.get('mode', 'harian')
    tanggal = request.args.get('tanggal', datetime.now(WIB).strftime('%Y-%m-%d'))

    cursor = conn.cursor()
    filter_query = '''
        SELECT p.id, b.nama, p.jumlah, p.harga_total, p.tanggal
        FROM pembelian p
        JOIN bahan b ON b.id = p.bahan_id
    '''
    params = []
    where_clause = ""

    if mode == 'harian':
        where_clause = "WHERE DATE(p.tanggal) = ?"
        params.append(tanggal)

    elif mode == 'mingguan':
        ref = datetime.strptime(tanggal, "%Y-%m-%d")
        start = (ref - timedelta(days=ref.weekday())).date()
        end = start + timedelta(days=6)
        where_clause = "WHERE DATE(p.tanggal) BETWEEN ? AND ?"
        params.extend([start.isoformat(), end.isoformat()])

    elif mode == 'bulanan':
        bulan = tanggal[:7]  # 'YYYY-MM'
        where_clause = "WHERE strftime('%Y-%m', p.tanggal) = ?"
        params.append(bulan)

    final_query = filter_query + " " + where_clause + " ORDER BY p.tanggal DESC"
    pembelian_data = cursor.execute(final_query, params).fetchall()

    total_pengeluaran = sum(p['harga_total'] for p in pembelian_data)

    bahan_list = conn.execute('SELECT * FROM bahan ORDER BY nama ASC').fetchall()

    conn.close()
    return render_template('pembelian.html',
                       bahan=bahan_list,
                       pembelian_template=pembelian_data,
                       total_pengeluaran=total_pengeluaran,
                       mode=mode,
                       tanggal=tanggal,
                       active_page='pembelian')



@app.route('/bahan/tambah', methods=['POST'])
def tambah_bahan():
    nama = request.form['nama']
    satuan = request.form['satuan']

    conn = get_db_connection()
    conn.execute('INSERT INTO bahan (nama, satuan) VALUES (?, ?)', (nama, satuan))
    conn.commit()
    conn.close()

    flash('✅ Bahan baru berhasil ditambahkan.')
    return redirect(url_for('pembelian'))



@app.route('/export/pembelian-ringkas-xlsx')
def export_pembelian_ringkas_xlsx():
    mode = request.args.get('mode', 'harian')
    tanggal_str = request.args.get('tanggal', datetime.now().strftime('%Y-%m-%d'))

    try:
        tanggal = datetime.strptime(tanggal_str, '%Y-%m-%d')
    except ValueError:
        return "❌ Format tanggal tidak valid. Gunakan YYYY-MM-DD", 400

    # Hitung rentang tanggal
    if mode == 'harian':
        tanggal_list = [tanggal.date()]
    elif mode == 'mingguan':
        start = (tanggal - timedelta(days=tanggal.weekday())).date()
        tanggal_list = [start + timedelta(days=i) for i in range(7)]
    elif mode == 'bulanan':
        start = tanggal.replace(day=1).date()
        next_month = (tanggal.replace(day=28) + timedelta(days=4)).replace(day=1)
        end = (next_month - timedelta(days=1)).date()
        tanggal_list = [start + timedelta(days=i) for i in range((end - start).days + 1)]
    else:
        return "❌ Mode tidak dikenali (gunakan harian, mingguan, bulanan)", 400

    conn = get_db_connection()
    cursor = conn.cursor()

    wb = Workbook()
    wb.remove(wb.active)

    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4B2C20")
    total_fill = PatternFill("solid", fgColor="D5BBA2")
    center_align = Alignment(horizontal="center")

    for tgl in tanggal_list:
        sheet = wb.create_sheet(title=tgl.strftime('%Y-%m-%d'))

        pembelian = cursor.execute('''
            SELECT b.nama, b.satuan, p.jumlah, p.harga_total
            FROM pembelian p
            JOIN bahan b ON b.id = p.bahan_id
            WHERE DATE(p.tanggal) = ?
            ORDER BY b.nama ASC
        ''', (tgl.isoformat(),)).fetchall()

        sheet.append(['PEMBELIAN BAHAN'])
        row = sheet.max_row
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        sheet.cell(row=row, column=1).font = Font(bold=True)
        sheet.cell(row=row, column=1).alignment = center_align

        sheet.append(['No', 'Nama Bahan', 'Jumlah', 'Satuan', 'Total (Rp)'])
        for col in range(1, 6):
            cell = sheet.cell(row=sheet.max_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        total_all = 0
        for idx, p in enumerate(pembelian, start=1):
            total_all += p['harga_total']
            sheet.append([idx, p['nama'], p['jumlah'], p['satuan'], p['harga_total']])

        sheet.append(['', '', '', 'TOTAL', total_all])
        for col in range(4, 6):
            cell = sheet.cell(row=sheet.max_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = total_fill

    conn.close()

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"pembelian_{mode}_{tanggal_str}_perhari.xlsx"
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={filename}"}
    )



def open_browser():
    webbrowser.open("http://localhost:5000")
    
if __name__ == '__main__':
    threading.Timer(1.5, open_browser).start()
    app.run(host='0.0.0.0', port=5000, debug=False)